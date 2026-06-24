from fastapi import FastAPI, HTTPException, Depends, Query, Response,status
from fastapi.middleware.cors import CORSMiddleware
from contextlib import asynccontextmanager
import httpx
import os
import io
from typing import Optional, List
from dotenv import load_dotenv
import json

from database import get_db, init_db, migrate_db
from auth import authenticate_user, get_current_session, require_superadmin, revoke_session
from models import (
    DeviceResponse, DeviceListResponse, ToggleStatusRequest, BulkToggleRequest,
    LoginRequest, LoginResponse, SyncResponse, DashboardStats,
    UpdateExpirationRequest, UpdateDeviceDetailsRequest, CustomFieldUpsert,
    MonthlyExpiration, ClientConfig, ClientConfigUpdate,
    SellerStats, InvoicePreview, RegisterAlertConfiguration
)
import crud

def _require_operator(session: dict):
    """Lanza 403 si el usuario es solo 'viewer'. Permite admin y operator."""
    if session.get("role") == "viewer" and not session.get("is_superadmin"):
        raise HTTPException(status_code=403, detail="Sin permisos para realizar esta accion")

load_dotenv()

FULLTRACK_BASE_URL = os.getenv("FULLTRACK_BASE_URL", "http://ws.fulltrack2.com")
FULLTRACK_APIKEY   = os.getenv("FULLTRACK_APIKEY", "")
FULLTRACK_SECRET   = os.getenv("FULLTRACK_SECRET", "")
ADMIN_USER         = os.getenv("ADMIN_USER", "admin")
ADMIN_PASSWORD     = os.getenv("ADMIN_PASSWORD", "controltrack2024")


def ft_url(path: str) -> str:
    return f"{FULLTRACK_BASE_URL}/{path}/apiKey/{FULLTRACK_APIKEY}/secretKey/{FULLTRACK_SECRET}"


async def do_sync():
    from database import get_pool
    import aiomysql as _aiomysql
    pool = await get_pool()
    total_synced = 0

    async with pool.acquire() as conn:
        async with conn.cursor(_aiomysql.DictCursor) as cur:
            await cur.execute("SELECT id, ft_apikey, ft_secretkey FROM tenants WHERE active = 1")
            tenants = await cur.fetchall()

    for tenant in tenants:
        tenant_id = tenant["id"]
        apikey    = tenant["ft_apikey"]
        secretkey = tenant["ft_secretkey"]

        def ft_url_tenant(path):
            return f"{FULLTRACK_BASE_URL}/{path}/apiKey/{apikey}/secretKey/{secretkey}"

        async with httpx.AsyncClient(timeout=30) as client:
            try:
                r_clients  = await client.get(ft_url_tenant("clients/all"))
                r_trackers = await client.get(ft_url_tenant("trackers/all"))
                r_vehicles = await client.get(ft_url_tenant("vehicles/all"))
                r_events   = await client.get(ft_url_tenant("events/all"))
                r_products = await client.get(ft_url_tenant("trackers/products"))
                r_workshop = await client.get(ft_url_tenant("workshop/list"))
            except httpx.RequestError as e:
                print(f"Error sync tenant {tenant_id}: {e}")
                continue

        async with pool.acquire() as conn:
            async with conn.cursor(_aiomysql.DictCursor) as cur:
                synced = await crud.sync_data(
                    cur,
                    r_clients.json().get("data", []),
                    r_trackers.json().get("data", []),
                    r_vehicles.json().get("data", []),
                    r_events.json().get("data", []),
                    r_products.json().get("data", []),
                    r_workshop.json().get("data", []),
                    tenant_id=tenant_id,
                )
                await conn.commit()
                total_synced += synced

    return total_synced


async def _whatsapp_scheduler_loop():
    """Corre el scheduler de WhatsApp cada hora. Envia notificaciones 1 y 3 dias antes del vencimiento."""
    import asyncio
    import aiomysql as _aiomysql
    from database import get_pool
    await asyncio.sleep(10)  # espera breve al arrancar
    while True:
        try:
            pool = await get_pool()
            async with pool.acquire() as conn:
                async with conn.cursor(_aiomysql.DictCursor) as cur:
                    from whatsapp import run_whatsapp_scheduler
                    result = await run_whatsapp_scheduler(cur)
                    if result["sent"] or result["failed"]:
                        print(f"[WhatsApp Scheduler] Enviados: {result['sent']} | Fallidos: {result['failed']} | Omitidos: {result['skipped']}")
        except Exception as e:
            print(f"[WhatsApp Scheduler] Error: {e}")
        await asyncio.sleep(3600)  # cada 1 hora


@asynccontextmanager
async def lifespan(app: FastAPI):
    import asyncio
    await init_db()
    await migrate_db()
    if FULLTRACK_APIKEY and FULLTRACK_SECRET:
        try:
            print("Sincronizando con Fulltrack...")
            synced = await do_sync()
            print(f"Sync completo: {synced} dispositivos")
        except Exception as e:
            print(f"Auto-sync fallo: {e}")
    # Lanzar scheduler de WhatsApp en background
    scheduler_task = asyncio.create_task(_whatsapp_scheduler_loop())
    yield
    scheduler_task.cancel()


app = FastAPI(title="ControlTrack API", version="3.0.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── Auth ─────────────────────────────────────────────────────────────────────

@app.post("/api/login", response_model=LoginResponse)
async def login(body: LoginRequest):
    if body.username == ADMIN_USER and body.password == ADMIN_PASSWORD:
        return LoginResponse(success=True, message="Autenticación exitosa")
    raise HTTPException(status_code=401, detail="Credenciales incorrectas")


# ─── Sync ─────────────────────────────────────────────────────────────────────

@app.post("/api/sync", response_model=SyncResponse)
async def sync_fulltrack():
    synced = await do_sync()
    return SyncResponse(success=True, message="Sincronización completa", synced_devices=synced)


# ─── Devices ──────────────────────────────────────────────────────────────────

@app.get("/api/devices", response_model=DeviceListResponse)
async def get_devices(
    search_client:         Optional[str] = None,
    search_imei:           Optional[str] = None,
    search_device:         Optional[str] = None,
    status_filter:         Optional[str] = None,
    expiring_days:         Optional[int] = Query(None),
    expire_from:           Optional[str] = Query(None),
    expire_to:             Optional[str] = Query(None),
    seller_filter:         Optional[str] = Query(None),
    installer_filter:      Optional[str] = Query(None),
    contract_type_filter:  Optional[str] = Query(None),
    tenant_id:             Optional[int] = Query(None),  # superadmin puede filtrar por tenant
    page:                  int = Query(1, ge=1),
    page_size:             int = Query(10, ge=1, le=1000),
    db=Depends(get_db),
    session=Depends(get_current_session)
):
    if session.get("is_superadmin"):
        # Superadmin: usa tenant_id del query param (None = ver todos los tenants)
        effective_tenant_id = tenant_id
    else:
        # Usuario de tenant: siempre ve solo sus datos, ignora query param
        effective_tenant_id = session.get("tenant_id")
    return await crud.get_devices(
        db, search_client, search_imei, search_device, status_filter,
        expiring_days, expire_from, expire_to,
        seller_filter, installer_filter, contract_type_filter,
        page, page_size, tenant_id=effective_tenant_id
    )


@app.get("/api/devices/{device_id}", response_model=DeviceResponse)
async def get_device(device_id: int, db=Depends(get_db)):
    device = await crud.get_device_by_id(db, device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Dispositivo no encontrado")
    return device


@app.put("/api/devices/{device_id}/expiration")
async def update_expiration(device_id: int, body: UpdateExpirationRequest,
                            db=Depends(get_db), session=Depends(get_current_session)):
    _require_operator(session)
    ok = await crud.update_expiration(db, device_id, body.expiration_date)
    if not ok:
        raise HTTPException(status_code=404, detail="Dispositivo no encontrado")
    return {"success": True, "message": "Fecha actualizada"}


@app.put("/api/devices/{device_id}/details")
async def update_device_details(device_id: int, body: UpdateDeviceDetailsRequest,
                                db=Depends(get_db), session=Depends(get_current_session)):
    """Actualiza tipo de contratación, vendedor, instalador, precio, RFC."""
    _require_operator(session)
    ok = await crud.update_device_details(db, device_id, body.model_dump(exclude_none=False))
    if not ok:
        raise HTTPException(status_code=404, detail="Dispositivo no encontrado")
    device = await crud.get_device_by_id(db, device_id)
    return {"success": True, "device": device}


@app.get("/api/devices/{device_id}/custom-fields")
async def get_device_custom_fields(device_id: int, db=Depends(get_db), session=Depends(get_current_session)):
    tenant_id = _effective_tenant(session, None)
    async with db.acquire() as conn:
        async with conn.cursor(aiomysql.DictCursor) as cur:
            q = "SELECT field_key, field_label, field_type, field_value FROM device_custom_fields WHERE device_id=%s"
            params = [device_id]
            if tenant_id is not None:
                q = """SELECT cf.field_key, cf.field_label, cf.field_type, cf.field_value
                       FROM device_custom_fields cf
                       JOIN devices d ON d.id = cf.device_id
                       WHERE cf.device_id=%s AND d.tenant_id=%s"""
                params.append(tenant_id)
            await cur.execute(q, params)
            return await cur.fetchall()

@app.put("/api/devices/{device_id}/custom-fields")
async def upsert_custom_field(device_id: int, body: CustomFieldUpsert,
                              db=Depends(get_db), session=Depends(get_current_session)):
    _require_operator(session)
    await crud.upsert_custom_field(
        db, device_id, body.field_key, body.field_label, body.field_type, body.field_value or ""
    )
    return {"success": True}


@app.delete("/api/devices/{device_id}/custom-fields/{field_key}")
async def delete_custom_field(device_id: int, field_key: str,
                              db=Depends(get_db), session=Depends(get_current_session)):
    _require_operator(session)
    await crud.delete_custom_field(db, device_id, field_key)
    return {"success": True}


@app.put("/api/devices/{device_id}/toggle")
async def toggle_device(device_id: int, body: ToggleStatusRequest,
                        db=Depends(get_db), session=Depends(get_current_session)):
    _require_operator(session)
    device = await crud.get_device_by_id(db, device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Dispositivo no encontrado")

    if body.deactivate:
        async with httpx.AsyncClient(timeout=15) as client:
            try:
                r = await client.get(ft_url("workshop/list"))
                workshop_data = r.json().get("data", [])
            except httpx.RequestError:
                workshop_data = []

        ins_id = vei_id = None
        for item in workshop_data:
            if str(item.get("ras_ras_id_aparelho", "")) == str(device["imei"]):
                ins_id = item.get("ras_ins_id")
                vei_id = item.get("ras_vei_id")
                break

        if ins_id:
            async with httpx.AsyncClient(timeout=15) as client:
                try:
                    await client.put(ft_url("workshop/uninstall"), json={"ras_ins_id": ins_id})
                except httpx.RequestError:
                    pass

        await crud.update_device_install_info(db, device_id, ins_id, vei_id)
        await crud.update_device_status(db, device_id, "deactivated")
    else:
        ins_id = device.get("ras_ins_id")
        vei_id = device.get("vehicle_id")
        imei   = device.get("imei")

        if not ins_id or not vei_id or not imei:
            raise HTTPException(
                status_code=400,
                detail="No se encontraron los datos de instalación. Sincroniza primero."
            )

        async with httpx.AsyncClient(timeout=15) as client:
            try:
                r = await client.post(
                    ft_url("workshop/install"),
                    json={
                        "ras_ins_id":          int(ins_id),
                        "ras_ras_id_aparelho": str(imei),
                        "ras_vei_id":          int(vei_id),
                    }
                )
                ft_resp = r.json()
                if not ft_resp.get("status"):
                    raise HTTPException(
                        status_code=502,
                        detail=f"Fulltrack: {ft_resp.get('message', 'Error al reinstalar')}"
                    )
            except httpx.RequestError as e:
                raise HTTPException(status_code=502, detail=f"Error conectando con Fulltrack: {str(e)}")

        await crud.update_device_status(db, device_id, "active")

    return {
        "success": True,
        "message": f"Dispositivo {'desinstalado' if body.deactivate else 'reinstalado'} correctamente"
    }


# ─── Bulk toggle ──────────────────────────────────────────────────────────────

@app.post("/api/devices/bulk-toggle")
async def bulk_toggle_devices(body: BulkToggleRequest,
                              db=Depends(get_db), session=Depends(get_current_session)):
    """Desactiva o activa varios dispositivos a la vez (solo actualiza status local)."""
    _require_operator(session)
    new_status = "deactivated" if body.deactivate else "active"
    count = await crud.bulk_update_status(db, body.device_ids, new_status)
    return {"success": True, "updated": count, "message": f"{count} dispositivos actualizados"}


# ─── Client account toggle ────────────────────────────────────────────────────

@app.put("/api/clients/{client_fulltrack_id}/toggle")
async def toggle_client(client_fulltrack_id: str, body: ToggleStatusRequest,
                        db=Depends(get_db), session=Depends(get_current_session)):
    _require_operator(session)
    endpoint = "clients/deactive" if body.deactivate else "clients/active"

    async with httpx.AsyncClient(timeout=15) as client:
        try:
            r = await client.put(
                ft_url(endpoint),
                json={"ras_cli_id": int(client_fulltrack_id)}
            )
            ft_resp = r.json()
            if not ft_resp.get("status"):
                raise HTTPException(
                    status_code=502,
                    detail=f"Fulltrack: {ft_resp.get('message', 'Error desconocido')}"
                )
        except httpx.RequestError as e:
            raise HTTPException(status_code=502, detail=f"Error conectando con Fulltrack: {str(e)}")

    new_status = "deactivated" if body.deactivate else "active"
    await crud.update_client_devices_status(db, client_fulltrack_id, new_status)
    return {"success": True, "message": f"Cuenta {'desactivada' if body.deactivate else 'activada'}"}


# ─── Client config (grace days, auto-deactivate) ──────────────────────────────

@app.get("/api/clients/{client_fulltrack_id}/config", response_model=ClientConfig)
async def get_client_config(client_fulltrack_id: str, db=Depends(get_db)):
    return await crud.get_client_config(db, client_fulltrack_id)

@app.put("/api/clients/{client_fulltrack_id}/config")
async def update_client_config(client_fulltrack_id: str, body: ClientConfigUpdate,
                               db=Depends(get_db), session=Depends(get_current_session)):
    _require_operator(session)
    existing        = await crud.get_client_config(db, client_fulltrack_id)
    grace_days      = body.grace_days      if body.grace_days      is not None else existing["grace_days"]
    auto_deactivate = body.auto_deactivate if body.auto_deactivate is not None else existing["auto_deactivate"]
    client_name     = body.client_name     or existing.get("client_name")
    # whatsapp_number: solo actualizar si fue enviado explicitamente en el request
    wa = body.whatsapp_number if "whatsapp_number" in body.model_fields_set else ...
    await crud.upsert_client_config(
        db, client_fulltrack_id, grace_days, auto_deactivate, client_name,
        whatsapp_number=wa,
    )
    return {"success": True}

@app.get("/api/client-configs")
async def get_all_client_configs(db=Depends(get_db)):
    return await crud.get_all_client_configs(db)


# ─── Scheduler ────────────────────────────────────────────────────────────────

@app.post("/api/scheduler/run")
async def run_scheduler(db=Depends(get_db)):
    """Ejecuta la lógica de desactivación automática con carencia."""
    result = await crud.run_auto_deactivation_scheduler(db)
    return {"success": True, **result}


# ─── Vehicles ─────────────────────────────────────────────────────────────────

@app.get("/api/vehicles/{vehicle_id}")
async def get_vehicle_detail(vehicle_id: str):
    async with httpx.AsyncClient(timeout=15) as client:
        try:
            r = await client.get(ft_url(f"vehicles/single/id/{vehicle_id}"))
            return r.json()
        except httpx.RequestError as e:
            raise HTTPException(status_code=502, detail=f"Error conectando con Fulltrack: {str(e)}")


# ─── Stats ────────────────────────────────────────────────────────────────────

@app.get("/api/stats", response_model=DashboardStats)
async def get_stats(db=Depends(get_db)):
    return await crud.get_stats(db)

@app.get("/api/stats/monthly", response_model=List[MonthlyExpiration])
async def get_monthly_stats(db=Depends(get_db)):
    return await crud.get_monthly_expirations(db)

@app.get("/api/stats/by-seller", response_model=List[SellerStats])
async def get_seller_stats(db=Depends(get_db)):
    return await crud.get_seller_stats(db)


# ─── Client devices & invoice ─────────────────────────────────────────────────

@app.get("/api/clients/{client_fulltrack_id}/devices", response_model=List[DeviceResponse])
async def get_client_devices(client_fulltrack_id: str, db=Depends(get_db)):
    return await crud.get_devices_by_client(db, client_fulltrack_id)

@app.get("/api/clients/{client_fulltrack_id}/invoice-preview", response_model=InvoicePreview)
async def get_invoice_preview(client_fulltrack_id: str, db=Depends(get_db)):
    return await crud.get_invoice_preview(db, client_fulltrack_id)


# ─── Export ───────────────────────────────────────────────────────────────────

@app.get("/api/export")
async def export_data(
    format:               str = Query("csv", regex="^(csv|xlsx|pdf)$"),
    status_filter:        Optional[str] = None,
    seller_filter:        Optional[str] = None,
    contract_type_filter: Optional[str] = None,
    expire_from:          Optional[str] = None,
    expire_to:            Optional[str] = None,
    expiring_days:        Optional[int] = None,
    db=Depends(get_db)
):
    rows = await crud.get_export_data(
        db, status_filter, seller_filter, contract_type_filter,
        expire_from, expire_to, expiring_days
    )

    headers_map = {
        "client_name":   "Cliente",
        "device_name":   "Vehículo",
        "plate":         "Placa",
        "imei":          "IMEI",
        "sim":           "SIM",
        "model":         "Modelo GPS",
        "contract_type": "Tipo contrato",
        "seller_name":   "Vendedor",
        "installer_name":"Instalador",
        "install_date":  "Fecha instalación",
        "registration_date": "Fecha alta",
        "expiration_date":   "Vencimiento",
        "days_until_expiration": "Días restantes",
        "monthly_price": "Precio mensual",
        "status":        "Estado",
        "rfc":           "RFC",
    }
    col_keys = list(headers_map.keys())
    col_labels = list(headers_map.values())

    STATUS_ES = {"active": "Activo", "expiring": "Por vencer", "expired": "Vencido", "deactivated": "Desactivado"}
    CONTRACT_ES = {"monthly": "Mensual", "quarterly": "Trimestral", "semiannual": "Semestral", "annual": "Anual", "lease": "Arrendamiento"}

    def fmt(row, key):
        v = row.get(key)
        if key == "status":
            return STATUS_ES.get(v, v or "")
        if key == "contract_type":
            return CONTRACT_ES.get(v, v or "")
        if v is None:
            return ""
        return str(v)

    if format == "csv":
        lines = [",".join(col_labels)]
        for row in rows:
            lines.append(",".join(f'"{fmt(row, k)}"' for k in col_keys))
        content = "\n".join(lines).encode("utf-8-sig")
        return Response(
            content=content,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=controltrack_export.csv"}
        )

    elif format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl no instalado. Ejecuta: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ControlTrack"

        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for col_idx, label in enumerate(col_labels, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r_idx, row in enumerate(rows, 2):
            for c_idx, key in enumerate(col_keys, 1):
                ws.cell(row=r_idx, column=c_idx, value=fmt(row, key))

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=controltrack_export.xlsx"}
        )

    elif format == "pdf":
        try:
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            raise HTTPException(status_code=500, detail="reportlab no instalado. Ejecuta: pip install reportlab")

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)
        styles = getSampleStyleSheet()

        # Solo columnas principales para PDF (ancho limitado)
        pdf_keys   = ["client_name","device_name","plate","imei","contract_type","seller_name","expiration_date","days_until_expiration","status"]
        pdf_labels = ["Cliente","Vehículo","Placa","IMEI","Contrato","Vendedor","Vencimiento","Días","Estado"]

        data = [pdf_labels]
        for row in rows:
            data.append([fmt(row, k) for k in pdf_keys])

        col_widths = [120, 100, 60, 110, 80, 90, 75, 40, 70]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTSIZE",   (0,0), (-1,0), 9),
            ("FONTSIZE",   (0,1), (-1,-1), 8),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f1f5f9")]),
            ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#cbd5e1")),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("LEFTPADDING",  (0,0), (-1,-1), 4),
            ("RIGHTPADDING", (0,0), (-1,-1), 4),
            ("TOPPADDING",   (0,0), (-1,-1), 3),
            ("BOTTOMPADDING",(0,0), (-1,-1), 3),
        ]))

        title = Paragraph("<b>ControlTrack — Reporte de dispositivos</b>", styles["Title"])
        from datetime import datetime as dt_now
        subtitle = Paragraph(f"Generado: {dt_now.now().strftime('%d/%m/%Y %H:%M')} | Total: {len(rows)} registros", styles["Normal"])
        doc.build([title, Spacer(1, 8), subtitle, Spacer(1, 12), t])
        buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=controltrack_export.pdf"}
        )

# ─── NEW: Auth multi-tenant ────────────────────────────────────────────────────

from models import LoginResponseV2, TenantCreate, TenantUpdate, TenantResponse, \
    UserCreate, UserUpdate, UserResponse, WhatsAppSendRequest
import crud_tenants
import whatsapp as wa

@app.post("/api/v2/login", response_model=LoginResponseV2)
async def login_v2(body: LoginRequest, db=Depends(get_db)):
    result = await authenticate_user(body.username, body.password, db)
    return LoginResponseV2(
        success=True,
        token=result["token"],
        role=result["role"],
        username=result["username"],
        tenant_id=result.get("tenant_id"),
        tenant_name=result.get("tenant_name"),
        is_superadmin=result.get("role") == "superadmin",
    )

@app.post("/api/v2/logout")
async def logout_v2(session: dict = Depends(get_current_session)):
    # token viene en el header; lo extraemos para revocarlo
    return {"success": True}


# ─── NEW: Tenants (solo super-admin) ──────────────────────────────────────────

@app.get("/api/tenants")
async def list_tenants(db=Depends(get_db), session=Depends(require_superadmin)):
    return await crud_tenants.get_tenants(db)

@app.post("/api/tenants", status_code=201)
async def create_tenant(body: TenantCreate, db=Depends(get_db), session=Depends(require_superadmin)):
    tenant_id = await crud_tenants.create_tenant(db, body.name, body.ft_apikey, body.ft_secretkey)
    return {"success": True, "id": tenant_id}

@app.put("/api/tenants/{tenant_id}")
async def update_tenant(tenant_id: int, body: TenantUpdate, db=Depends(get_db), session=Depends(require_superadmin)):
    await crud_tenants.update_tenant(db, tenant_id, body.model_dump(exclude_none=True))
    return {"success": True}

@app.delete("/api/tenants/{tenant_id}")
async def delete_tenant(tenant_id: int, db=Depends(get_db), session=Depends(require_superadmin)):
    await crud_tenants.delete_tenant(db, tenant_id)
    return {"success": True}

@app.post("/api/tenants/{tenant_id}/sync")
async def sync_tenant(tenant_id: int, db=Depends(get_db), session=Depends(require_superadmin)):
    """Sincroniza Fulltrack usando las credenciales del tenant."""
    tenant = await crud_tenants.get_tenant(db, tenant_id)
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant no encontrado")

    ft_base = os.getenv("FULLTRACK_BASE_URL", "http://ws.fulltrack2.com")
    def t_url(path): return f"{ft_base}/{path}/apiKey/{tenant['ft_apikey']}/secretKey/{tenant['ft_secretkey']}"

    async with httpx.AsyncClient(timeout=30) as client:
        try:
            r_clients  = await client.get(t_url("clients/all"))
            r_trackers = await client.get(t_url("trackers/all"))
            r_vehicles = await client.get(t_url("vehicles/all"))
            r_events   = await client.get(t_url("events/all"))
            r_products = await client.get(t_url("trackers/products"))
            r_workshop = await client.get(t_url("workshop/list"))
        except httpx.RequestError as e:
            raise HTTPException(status_code=502, detail=str(e))

    from database import get_pool
    import aiomysql as _aiomysql
    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.cursor(_aiomysql.DictCursor) as cur:
            synced = await crud.sync_data(
                cur,
                r_clients.json().get("data", []),
                r_trackers.json().get("data", []),
                r_vehicles.json().get("data", []),
                r_events.json().get("data", []),
                r_products.json().get("data", []),
                r_workshop.json().get("data", []),
                tenant_id=tenant_id,  # ✅ Bug 1 corrigido: tenant_id agora é passado
            )
        await conn.commit()
    return {"success": True, "synced": synced}


# ─── NEW: Users ────────────────────────────────────────────────────────────────

@app.get("/api/users")
async def list_users(tenant_id: Optional[int] = None, db=Depends(get_db),
                     session=Depends(get_current_session)):
    # Superadmin ve todos; tenant-admin solo los suyos
    if session.get("is_superadmin"):
        return await crud_tenants.get_users(db, tenant_id)
    return await crud_tenants.get_users(db, session["tenant_id"])

@app.post("/api/users", status_code=201)
async def create_user(body: UserCreate, db=Depends(get_db), session=Depends(get_current_session)):
    if not session.get("is_superadmin") and session.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Sin permisos para crear usuarios")
    user_id = await crud_tenants.create_user(
        db, body.tenant_id, body.username, body.password, body.full_name, body.role
    )
    return {"success": True, "id": user_id}

@app.put("/api/users/{user_id}")
async def update_user(user_id: int, body: UserUpdate, db=Depends(get_db),
                      session=Depends(get_current_session)):
    if not session.get("is_superadmin") and session.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Sin permisos")
    await crud_tenants.update_user(db, user_id, body.model_dump(exclude_none=True))
    return {"success": True}

@app.delete("/api/users/{user_id}")
async def delete_user(user_id: int, db=Depends(get_db), session=Depends(get_current_session)):
    if not session.get("is_superadmin") and session.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Sin permisos")
    await crud_tenants.delete_user(db, user_id)
    return {"success": True}


# ─── NEW: WhatsApp ─────────────────────────────────────────────────────────────

@app.post("/api/whatsapp/send")
async def send_whatsapp(body: WhatsAppSendRequest, db=Depends(get_db),
                        session=Depends(get_current_session)):
    """Envía recordatorio manual de renovación a un número de WhatsApp."""
    _require_operator(session)
    tenant_id = None if session.get("is_superadmin") else session.get("tenant_id")
    result = await wa.send_renewal_notification(
        db, body.device_ids, body.whatsapp_number, tenant_id
    )
    if result.get("failed") and not result.get("sent"):
        raise HTTPException(status_code=502, detail=result.get("error", "Error al enviar"))
    return {"success": True, **result}

@app.post("/api/whatsapp/scheduler")
async def run_whatsapp_scheduler(db=Depends(get_db), session=Depends(get_current_session)):
    """Ejecuta el scheduler de notificaciones automáticas."""
    tenant_id = None if session.get("is_superadmin") else session.get("tenant_id")
    result = await wa.run_whatsapp_scheduler(db, tenant_id)
    return {"success": True, **result}

@app.get("/api/whatsapp/history")
async def get_whatsapp_history(db=Depends(get_db), session=Depends(get_current_session)):
    tenant_id = None if session.get("is_superadmin") else session.get("tenant_id")
    return await crud_tenants.get_notification_history(db, tenant_id)



@app.post("/api/monitoring")
async def register_alert_configuration(body: RegisterAlertConfiguration,db=Depends(get_db), session=Depends(get_current_session)):
    tenant_id = None if session.get("is_superadmin") else session.get("tenant_id")

    if body.notification_channel in ("whatsapp", "both") and not body.phone_number:
        raise HTTPException(400, "phone_number es requerido para el canal seleccionado")
    if body.notification_channel in ("email", "both") and not body.email:
        raise HTTPException(400, "email es requerido para el canal seleccionado")
    body = body.model_dump()

    alert_config_response = await crud_tenants.create_alert_configuration(db,tenant_id,body)

    if alert_config_response:
        return {"success":True}
    else:
        raise HTTPException(status_code=500, detail="Error inserting alert configuration")

    
    
@app.get("/api/monitoring")
async def get_alert_configuration(db=Depends(get_db), session=Depends(get_current_session)):
    tenant_id = None if session.get("is_superadmin") else session.get("tenant_id")
    config_response = await crud_tenants.get_alert_configuration(db,tenant_id)
    if config_response:
        data = json.dumps(config_response)
        return {"data": config_response}
    
    else:
        return Response(status_code=status.HTTP_204_NO_CONTENT)
